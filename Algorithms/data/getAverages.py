from openpyxl import Workbook # Import library to work with Microsoft Excel files
wb = Workbook()
ws1 = wb.active
ws1.title = "Size Averages" # Rename first sheet 
ws2 = wb.create_sheet(title="Board Averages") # Creat second sheet with the specified title 

def main():
    """Gather all the raw data from the individual files and compile averages into both sheets"""
    path = "./raw/"
    for a in ("f", "m"):
        for s in range(2, 6):
            sTotal = [0, 0, 0, 0] # Totals for the board size 
            for b in range(1, 501):
                average = getBoardAverage(path, a, s, b)
                for d in range(4):
                    sTotal[d] += average[d]
            for d in range(4):
                sTotal[d] = float(sTotal[d]/500) # Averages for the size 
                if a == "m":
                    ws1.cell(row=s-1, column=6+d, value=sTotal[d]) # Write data to the Excel sheet 
            if a == "f":
                ws1.append([s] + sTotal) # Write data to the Excel sheet 
    wb.save("processed.xlsx") # Save the Excel file 

def getBoardAverage(path, a, s, b):
    """Gather the average for each metric from each of the 5 trials of a given board"""
    totals = [0, 0, 0, 0] # Totals for the given board 
    for t in range(1, 6):
        with open(f"{path}{'.'.join((a, str(s), str(b)))}.{t}.txt", "r") as f:
            lines = f.readlines()
            for l in range(4):
                totals[l] += int(lines[l])
    for d in range(4):
        totals[d] = totals[d]/5 # Averages for the given board 
        if a == "m":
            ws2.cell(row=((s-2)*500)+b, column=6+d, value=totals[d]) # Write data to Excel sheet 
    if a == "f":
        ws2.append([f"{s}.{b}"] + totals) # Write data to Excel sheet 
    return totals

main()
